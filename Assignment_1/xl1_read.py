import openpyxl
wb = openpyxl.load_workbook('Assignment_1_Data_and_Template.xlsx')
print (type(wb))
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Data')
print (type(sheet))
print (sheet['A1'].value)
print (sheet.max_row)
print (sheet.max_column)

