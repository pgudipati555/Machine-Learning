import openpyxl
import os
import struct
import matplotlib as plt
from array import array as pyarray
from pylab import *
from numpy import *
import numpy.linalg as la
import statistics
import math

#************************** Function for Getting BIN Number for a given height ***************
def get_bin_num (ht_i,sp_i, ht_max, ht_min, sp_max, sp_min, bin_cnt):
    (r,c)=(0,0)
    r = round ((bin_cnt-1) * ((ht_i-ht_min)/(ht_max-ht_min)))
    c = round ((bin_cnt-1) * ((sp_i-sp_min)/(sp_max-sp_min)))
    return (r,c)

def get_pd (x, mu, cov):
    output2 = ""    
    diff = subtract(x,mu)
    diff_trans = diff.transpose()
    cov_det = la.det(cov)
    cov_inv = la.inv(cov)

    tmpa = dot(diff, cov_inv)
    tmpb = dot(tmpa, diff_trans)
    
    tmp1 = (1 / (2 * math.pi * math.sqrt(cov_det)) )
    tmp2 = (-0.5 * tmpb )
    tmp3 = math.exp (tmp2)
    prob_den = tmp1 * tmp3
    return(prob_den)

def bayesian (Nm, pm, Nf, pf):
    temp1 = (Nm*pm) /(Nm+Nf)
    temp2 = (Nf*pf)/(Nm+Nf)
    Pf = temp2 / (temp1+temp2)
    return (Pf)

def bayesian2 (Nm, pm, Nf, pf):
    Pf = (Nf*pf) / ((Nf*pf) + (Nm*pm) )
    return (Pf)    


def load_mnist(dataset="training", digits=range(10), path='C:\\Users\\pgudipati\\Me-Pras-Cloud\\Technical-Docs\\MachineLearning\\Assignment_3'):
    
    """
    Adapted from: http://cvxopt.org/applications/svm/index.html?highlight=mnist
    """

    if dataset == "training":
        fname_img = os.path.join(path, 'train-images.idx3-ubyte')
        fname_lbl = os.path.join(path, 'train-labels.idx1-ubyte')
    elif dataset == "testing":
        fname_img = os.path.join(path, 't10k-images.idx')
        fname_lbl = os.path.join(path, 't10k-labels.idx')
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    flbl = open(fname_lbl, 'rb')
    magic_nr, size = struct.unpack(">II", flbl.read(8))
    lbl = pyarray("b", flbl.read())
    flbl.close()

    fimg = open(fname_img, 'rb')
    magic_nr, size, rows, cols = struct.unpack(">IIII", fimg.read(16))
    img = pyarray("B", fimg.read())
    fimg.close()

    ind = [ k for k in range(size) if lbl[k] in digits ]
    N = len(ind)

    images = zeros((N, rows, cols), dtype=uint8)
    labels = zeros((N, 1), dtype=int8)
    for i in range(len(ind)):
        images[i] = array(img[ ind[i]*rows*cols : (ind[i]+1)*rows*cols ]).reshape((rows, cols))
        labels[i] = lbl[ind[i]]

    return images, labels

#************* End of Functions Definitions  ***************

#****************************************************************************************
#************************************ Main Program  **********************************
#****************************************************************************************       
output = "START Execution"
output2 = ""
print (output2.center(60, "*"))
print (output.center(60, "*"))
print (output2.center(60, "*"))

labeln = 6
labelp = 7

images, labels = load_mnist('training', digits=[labeln,labelp])

## converting from NX28X28 array into NX784 array
flatimages = list()
for i in images:
    flatimages.append(i.ravel())
X = np.asarray(flatimages)

print("Check shape of matrix", X.shape)
print("Check Mins and Max Values",np.amin(X),np.amax(X))

#****************************************************************************************
#******************** Check training vector by plotting image *******************
#****************************************************************************************       
#print("\nCheck training vector by plotting image \n")
#plt.imshow(X[20].reshape(28, 28),interpolation='None', cmap=cm.gray)
#show()


#************************************************************************************
print("1. Calculate Mu: Mean of X")
Mu = mean(X,axis=0)
print("2. Calculate Z: X-Mu")
Z = X-Mu
print("3. Calculate C: Coverence Matrix of Z")
C = cov(Z,rowvar= False)
print("4. Calculate V: Eigen Vectors of C")
[Lambda,V]=la.eigh(C)
print("5.1 *** reverse the order of eigenvectors and eigenvalues so that they are ordered in decreasing order of importance ")
print("5.2 *** Also set V = V.T ")
Lambda=flipud(Lambda)
V=flipud(V.T)
print("5.3 *** Reduce V diemntsions to 2. Now V will be 2 X 784")
V = V[0:2, :]
print("6. Calculate P: Dot Product of Z and V.Transpose")
P=dot(Z,V.T);
print("7. Calculate R")
R=dot(P,V)
print("8. Calculate diffrz : R-Z : All the values should be Very small")
diffrz = R-Z
print("9. Calculate Xrec2: Reconstruct X using 2D Principal Components")
Xrec2= R + Mu

print("10. Draw Scatter Plot")
T = labels
cols=zeros((alen(T),4));
for y in (range (alen(T)) ):
    if (T[y] == labeln):
	    cols[y]=[1,0,0,0.25];
    elif (T[y] == labelp):
	    cols[y]=[0,1,0,0.25];

randomorder=permutation(arange(alen(T))); #Don't worry about this stuff. Just makes a pretty picture
fig = figure()
ax = fig.add_subplot(111, facecolor='black')
ax.scatter(P[randomorder,1],P[randomorder,0],s=5,linewidths=0,facecolors=cols[randomorder,:],marker="o");
ax.set_aspect('equal');
gca().invert_yaxis();
show()

#*********************************************
Neg = []
Pos = []
Neg_And_Pos = []
Bin_count = 25

for y in (range (alen(T)) ):
    Neg_And_Pos.append([P[y][0], P[y][1]])
    if (T[y] == labeln):
	    Neg.append([P[y][0], P[y][1]])
    elif (T[y] == labelp):
	    Pos.append([P[y][0], P[y][1]])

Na = array(Neg)
Pa = array(Pos)
NPa = array (Neg_And_Pos)

#******* Create Neg and Pos Histograms *********************************************

Male = Neg
Female = Pos
Male_And_Female = Neg_And_Pos

M = Na
F = Pa
MF = NPa

(Nm,tmp) = np.shape(M)
(Nf,tmp) = np.shape(F)

(ht_max, sp_max) = np.amax(MF, axis=0)
(ht_min, sp_min) = np.amin(MF, axis=0)

print ("Size of/Rows in M/Negative Class: ", Nm,"\nSize of/Rows in F/Positive Class:", Nf) 

if  ( (Nm+Nf == alen(P)) & (Nm+Nf == alen(T)) ):
    print ("1. Number of Feature Vectors in the Training Set = Sum of Rowns in Male and Female Arrays- Check PASS")
else:
    print ("1. Number of Feature Vectors in the Training Set != Sum of Rowns in Male and Female Arrays- Check FAIL")

output = ""
print (output.center(30, "*"))
print ("Number of Male/Negative Samples: ",Nm)
print ("Number of Female/Positive Samples: ",Nf)
print ("Overall Max. Height/1st D : ",ht_max)
print ("Overall Min. Height/1st D: ",ht_min)
print ("Overall Max. Hand Span/2nd D: ",sp_max)
print ("Overall Min. Hand Span/2nd D: ",sp_min)
print ("Bin Count: ",Bin_count)
print (output.center(30, "*"))

# Create Male and Female Histograms *********************************************
output = "Create Male Histogram: "
print (output.center(100, "*"))

Hm = np.zeros((Bin_count,Bin_count), int)
Hf = np.zeros((Bin_count,Bin_count), int)

for (ht_i,sp_i) in (Male):
    (r,c) = get_bin_num(ht_i,sp_i, ht_max, ht_min, sp_max, sp_min,Bin_count)
    Hm[(int(r))][(int(c))]=Hm[(int(r))][(int(c))]+1
    #Hm[r][c]=Hm[r][c]+1

for (ht_i,sp_i) in (Female):
    (r,c) = get_bin_num(ht_i,sp_i, ht_max, ht_min, sp_max, sp_min,Bin_count)
    Hf[(int(r))][(int(c))]=Hf[(int(r))][(int(c))]+1

if ((Nm+Nf) == (np.sum(Hm)+np.sum(Hf))):
    print ("Number of Feature Vectors in the Training Set = Sum of Hm+ Sum of Hf : Check PASS")
else:
    print ("Number of Feature Vectors in the Training Set != Sum of Hm+ Sum of Hf : Check FAIL")


#***************** Using Bayesian Classifier  ******************************
output = "Results using Bayesian Classifier"
print (output.center(100, "~"))
MuF = mean(F, axis=0)
MuM = mean(M, axis=0)
CovF=cov(F,rowvar=False)
CovM=cov(M,rowvar=False)
print("MuF :", MuF)
print("MuM :", MuM)
print("CovF :\n", CovF)
print("CovM :\n", CovM)

#*********************************
#****************************************************************************************
#************************************ Start of Testing **********************************
#****************************************************************************************
         
output = "Start of Testing "
output2 = ""
print (output2.center(100, "*"))
print (output.center(100, "*"))
print (output2.center(100, "*"))
tcs = [P[0], P[1], P[2], P[3], P[4], P[5], P[449]]
expected_results = [T[0][0],T[1][0],T[2][0],T[3][0],T[4][0],T[5][0], T[449][0], ]

#***************** Using Histograms  ******************************
output = "Results using Histogram Classifier"
print (output.center(100, "~"))

i = 0
exp_class = ""
for (ht_i,sp_i) in (tcs):
    (r,c) = get_bin_num(ht_i,sp_i, ht_max, ht_min, sp_max, sp_min,Bin_count)
    male_cnt = Hm[(int(r))][(int(c))]
    female_cnt = Hf[(int(r))][(int(c))]
    if ((female_cnt+male_cnt) > 0):
        p_f = (female_cnt/(female_cnt+male_cnt))
    else:
        p_f = "Undefined"        
    print ("Height:", ht_i, "hand Span",sp_i, "\nbin number:", r,c, "Male/Neg Count:", male_cnt, "female/POS count:", female_cnt, "Probability of being female/POS: ", p_f, sep ="\t")
    if (expected_results[i] == labeln ):
        exp_class = "Negative"
    elif  (expected_results[i] == labelp ):
        exp_class = "Positive"
    print ("Expected Result:", expected_results[i], "\t", exp_class, "\n\n"); i=i+1;

#***************** Using Bayesian Classifier  ******************************
output = "Results using Bayesian Classifier"
print (output.center(100, "~"))

i=0
for (ht_i,sp_i) in (tcs):
    pd_m = get_pd (np.array([ht_i,sp_i]), MuM, CovM)
    pd_f = get_pd (np.array([ht_i,sp_i]), MuF, CovF)
    Pf1 = bayesian(Nm, pd_m, Nf, pd_f)
    Pf2 = bayesian2(Nm, pd_m, Nf, pd_f)
    print ("Height:", ht_i, "Hand Span:", sp_i, "\nPD_Male/NEG:", pd_m, "PD_female/POS:", pd_f, "Pf1/P_Pos1:", Pf1, "Pf2:P_Pos2", Pf2, sep ="\t")
    if (expected_results[i] == labeln ):
        exp_class = "Negative"
    elif  (expected_results[i] == labelp ):
        exp_class = "Positive"
    print ("Expected Result:", expected_results[i], "\t", exp_class, "\n\n"); i=i+1;

#***************** Calculate Training Accuracy ******************************
i = 0
correct_count = 0
incorrect_count = 0
for (ht_i,sp_i) in (P):
    (r,c) = get_bin_num(ht_i,sp_i, ht_max, ht_min, sp_max, sp_min,Bin_count)
    male_cnt = Hm[(int(r))][(int(c))]
    female_cnt = Hf[(int(r))][(int(c))]
    if ((female_cnt+male_cnt) > 0):
        p_f = (female_cnt/(female_cnt+male_cnt))
    else:
        p_f = "Undefined"

    if ((p_f > 0.5) and (T[i][0] == labelp)):
        correct_count = correct_count + 1
    elif ((p_f < 0.5) and (T[i][0] == labeln)):
        correct_count = correct_count + 1
    else:
        incorrect_count = incorrect_count + 1
    i=i+1

print ("Training Accuracy using Histogram# correct_count:", correct_count, "incorrect_count:", incorrect_count)
print ("Training Accuracy using Histogram# Percentage ", (correct_count/(incorrect_count+correct_count)) )
#
i = 0
correct_count = 0
incorrect_count = 0
for (ht_i,sp_i) in (P):
    pd_m = get_pd (np.array([ht_i,sp_i]), MuM, CovM)
    pd_f = get_pd (np.array([ht_i,sp_i]), MuF, CovF)
    Pf1 = bayesian(Nm, pd_m, Nf, pd_f)
    if ((Pf1 > 0.5) and (T[i][0] == labelp)):
        correct_count = correct_count + 1
    elif ((Pf1 < 0.5) and (T[i][0] == labeln)):
        correct_count = correct_count + 1
    else:
        incorrect_count = incorrect_count + 1
    i=i+1


print ("Training Accuracy using Bayesian# correct_count:", correct_count, "incorrect_count:", incorrect_count)
print ("Training Accuracy using Bayesian# Percentage ", (correct_count/(incorrect_count+correct_count)) )



output = "Writing Results to the XL sheet"
print (output.center(100, "^"))

wb = openpyxl.load_workbook('Assignment_3_ results-ashok.xlsx')
sheet = wb.get_sheet_by_name('Results')

print ("Write Mu results to XL")
for i in (range (alen(Mu))):
    sheet.cell(row=2, column=(2+i)).value = Mu[i]

print ("Write Eigen Vector 1 results to XL")
for i in (range (alen(V[0]))):
    sheet.cell(row=3, column=(2+i)).value = V[0][i]

print ("Write Eigen Vector 2 results to XL")
for i in (range (alen(V[1]))):
    sheet.cell(row=4, column=(2+i)).value = V[1][i]

print ("Write Positive Class Histogram to XL")
for i in (range(Bin_count)):
    for j in (range(Bin_count)):
        sheet.cell(row=(20+i), column=(2+j)).value = Hf[i][j]

print ("Write Negative Class Histogram to XL")
for i in (range(Bin_count)):
    for j in (range(Bin_count)):
        sheet.cell(row=(46+i), column=(2+j)).value = Hm[i][j]        

###########################################
print ("***** Write Xp[3] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=74, column=(2+i)).value = X[3][i]

print ("***** Write Zp[3] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=75, column=(2+i)).value = Z[3][i]

print ("***** Write Pp[3] XL")
for i in (range (len(P[3]))):
    sheet.cell(row=76, column=(2+i)).value = P[3][i]

print ("***** Write Rp[3] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=77, column=(2+i)).value = R[3][i]

print ("***** Write Xrec2p[3] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=78, column=(2+i)).value = Xrec2[3][i]

###########################################
print ("***** Write Xp[0] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=80, column=(2+i)).value = X[0][i]

print ("***** Write Zp[0] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=81, column=(2+i)).value = Z[0][i]

print ("***** Write Pp[0] XL")
for i in (range (len(P[0]))):
    sheet.cell(row=82, column=(2+i)).value = P[0][i]

print ("***** Write Rp[0] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=83, column=(2+i)).value = R[0][i]

print ("***** Write Xrec2p[0] XL")
for i in (range (alen(Mu))):
    sheet.cell(row=84, column=(2+i)).value = Xrec2[0][i]

#    
wb.save('Assignment_3_ results-ashok.xlsx')
