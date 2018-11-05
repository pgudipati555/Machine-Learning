import openpyxl
import numpy as np
import statistics
import math

wb = openpyxl.load_workbook('Assignment_4_Data_and_Template.xlsx')
sheet = wb.get_sheet_by_name('Training Data')
last_row_number = sheet.max_row
rows_cnt = last_row_number - 1
print ("Max row Number in XL sheet=", last_row_number)
print ("Total Number of Rows in the Training set =", rows_cnt)

###1. Initialize Data structures.
X = np.zeros((rows_cnt,15),dtype=int)
Xa = np.zeros((rows_cnt,16),dtype=int)
T1 = np.zeros((rows_cnt,1),dtype=int)
T2 = np.zeros((rows_cnt,6),dtype=int)
T2b = np.zeros((rows_cnt,1),dtype=int)
BinConfM = np.zeros((2,2),dtype=int)
SixConfM = np.zeros((6,6),dtype=int)

x = np.zeros((50,15),dtype=int)
xa = np.zeros((50,16),dtype=int)
t1 = np.zeros((50,1),dtype=int)
t2 = np.zeros((50,1),dtype=int)

#2. Read XL sheet
n =0
for i in (range(2, (last_row_number+1) )):
#for i in (range(2, 11 )):
    tmp1 = [ sheet["A"+str(i)].value, sheet["B"+str(i)].value, sheet["C"+str(i)].value, sheet["D"+str(i)].value,
             sheet["E"+str(i)].value, sheet["F"+str(i)].value, sheet["G"+str(i)].value, sheet["H"+str(i)].value,
             sheet["I"+str(i)].value, sheet["J"+str(i)].value, sheet["K"+str(i)].value, sheet["L"+str(i)].value,
             sheet["M"+str(i)].value, sheet["N"+str(i)].value, sheet["O"+str(i)].value, 
             ]
    tmp2 = [1]
    tmp2.extend(tmp1)
    X[n] = tmp1
    Xa[n] = tmp2 
    T1[n] = [sheet["P"+str(i)].value]
    tmp4 = [-1, -1, -1, -1, -1, -1]
    tmp4[sheet["Q"+str(i)].value] = 1
    T2b[n] = sheet["Q"+str(i)].value
    T2[n] = tmp4
    n = n +1
#3.
Xapi = np.linalg.pinv(Xa)
W1 = np.dot(Xapi, T1)
W2 = np.dot(Xapi, T2)


print ("Xa.shape:", Xa.shape)
print ("Xapi.shape:", Xapi.shape)
print ("W1.shape:", W1.shape)
print ("W2.shape:", W2.shape)

#***************** Queries  ********************

sheet2 = wb.get_sheet_by_name('To be classified')

n =0
for i in (range(5, 55)):
    tmp1 = [ sheet2["A"+str(i)].value, sheet2["B"+str(i)].value, sheet2["C"+str(i)].value, sheet2["D"+str(i)].value,
             sheet2["E"+str(i)].value, sheet2["F"+str(i)].value, sheet2["G"+str(i)].value, sheet2["H"+str(i)].value,
             sheet2["I"+str(i)].value, sheet2["J"+str(i)].value, sheet2["K"+str(i)].value, sheet2["L"+str(i)].value,
             sheet2["M"+str(i)].value, sheet2["N"+str(i)].value, sheet2["O"+str(i)].value, 
             ]
    tmp2 = [1]
    tmp2.extend(tmp1)
    x[n] = tmp1
    xa[n] = tmp2
    t1_res = np.dot(xa[n], W1)
    if (t1_res >0):
        t1[n] =1
    else:
        t1[n] = -1
        
    t2_temp = list(np.dot(xa[n], W2))
    t2[n] = t2_temp.index( max(t2_temp)) 
    n = n+1

#**************** Performance
print ("*** Calculate Binary Confusion Matrix *****")
for i in (range(rows_cnt)):
#for i in (range(10)):
    t1_res = np.dot(Xa[i], W1)
    if   ((T1[i] < 0) and (t1_res < 0)):
        BinConfM[0][0] =  BinConfM[0][0] + 1
    elif ((T1[i] < 0) and (t1_res > 0)):
        BinConfM[0][1] =  BinConfM[0][1] + 1
    elif ((T1[i] > 0) and (t1_res < 0)):
        BinConfM[1][0] =  BinConfM[1][0] + 1
    elif ((T1[i] > 0) and (t1_res > 0)):
        BinConfM[1][1] =  BinConfM[1][1] + 1

    #n = n+1
    #print ("Xa[i]:", Xa[i], "\t", "t1_res", t1_res, "\t", "T1[i]:",  T1[i])

print ("Binary Confusion Matrix:BinConfM: \n", BinConfM)

TN = BinConfM[0][0]
TP = BinConfM[1][1]
FN = BinConfM[1][0]
FP = BinConfM[0][1]

Accuracy = (TP + TN) / (TP+TN+FN+FP)
Sensitivity = TP / (TP+FN)
Specificity = TN / (FP+TN)
PPVb = TP / (FP+TP)

print ("Accuracy:", Accuracy)
print ("Sensitivity:", Sensitivity)
print ("Specificity:", Specificity)
print ("PPVb:", PPVb)


#**************** Performance : 6C Confusion Matrix
print ("*** Calculate 6 Class Confusion Matrix *****")
#for i in (range(30)):
for i in (range(rows_cnt)):

    t2_temp = list(np.dot(Xa[i], W2))
    #print ("t2_temp:", t2_temp)
    t2_class = t2_temp.index( max(t2_temp))


    if (T2b[i]==0):
        if (t2_class == 0):
            SixConfM[0][0] =  SixConfM[0][0] + 1
        elif (t2_class == 1):
            SixConfM[0][1] =  SixConfM[0][1] + 1
        elif (t2_class == 2):
            SixConfM[0][2] =  SixConfM[0][2] + 1
        elif (t2_class == 3):
            SixConfM[0][3] =  SixConfM[0][3] + 1
        elif (t2_class == 4):
            SixConfM[0][4] =  SixConfM[0][4] + 1
        elif (t2_class == 5):
            SixConfM[0][5] =  SixConfM[0][5] + 1
    elif (T2b[i]==1):
        if (t2_class == 0):
            SixConfM[1][0] =  SixConfM[1][0] + 1
        elif (t2_class == 1):
            SixConfM[1][1] =  SixConfM[1][1] + 1
        elif (t2_class == 2):
            SixConfM[1][2] =  SixConfM[1][2] + 1
        elif (t2_class == 3):
            SixConfM[1][3] =  SixConfM[1][3] + 1
        elif (t2_class == 4):
            SixConfM[1][4] =  SixConfM[1][4] + 1
        elif (t2_class == 5):
            SixConfM[1][5] =  SixConfM[1][5] + 1
    elif (T2b[i]==2):
        if (t2_class == 0):
            SixConfM[2][0] =  SixConfM[2][0] + 1
        elif (t2_class == 1):
            SixConfM[2][1] =  SixConfM[2][1] + 1
        elif (t2_class == 2):
            SixConfM[2][2] =  SixConfM[2][2] + 1
        elif (t2_class == 3):
            SixConfM[2][3] =  SixConfM[2][3] + 1
        elif (t2_class == 4):
            SixConfM[2][4] =  SixConfM[2][4] + 1
        elif (t2_class == 5):
            SixConfM[2][5] =  SixConfM[2][5] + 1
    elif (T2b[i]==3):
        if (t2_class == 0):
            SixConfM[3][0] =  SixConfM[3][0] + 1
        elif (t2_class == 1):
            SixConfM[3][1] =  SixConfM[3][1] + 1
        elif (t2_class == 2):
            SixConfM[3][2] =  SixConfM[3][2] + 1
        elif (t2_class == 3):
            SixConfM[3][3] =  SixConfM[3][3] + 1
        elif (t2_class == 4):
            SixConfM[3][4] =  SixConfM[3][4] + 1
        elif (t2_class == 5):
            SixConfM[3][5] =  SixConfM[3][5] + 1
    elif (T2b[i]==4):
        if (t2_class == 0):
            SixConfM[4][0] =  SixConfM[4][0] + 1
        elif (t2_class == 1):
            SixConfM[4][1] =  SixConfM[4][1] + 1
        elif (t2_class == 2):
            SixConfM[4][2] =  SixConfM[4][2] + 1
        elif (t2_class == 3):
            SixConfM[4][3] =  SixConfM[4][3] + 1
        elif (t2_class == 4):
            SixConfM[4][4] =  SixConfM[4][4] + 1
        elif (t2_class == 5):
            SixConfM[4][5] =  SixConfM[4][5] + 1
    elif (T2b[i]==5):
        if (t2_class == 0):
            SixConfM[5][0] =  SixConfM[5][0] + 1
        elif (t2_class == 1):
            SixConfM[5][1] =  SixConfM[5][1] + 1
        elif (t2_class == 2):
            SixConfM[5][2] =  SixConfM[5][2] + 1
        elif (t2_class == 3):
            SixConfM[5][3] =  SixConfM[5][3] + 1
        elif (t2_class == 4):
            SixConfM[5][4] =  SixConfM[5][4] + 1
        elif (t2_class == 5):
            SixConfM[5][5] =  SixConfM[5][5] + 1

    #print ("Xa[i]:", Xa[i], "\t", "t2_class", t2_class, "\t", "T2b[i]:",  T2b[i])
print ("Six Class Confusion Matrix:SixConfM: \n", SixConfM)

##### Writing Results to XLS
#output = "Writing Results to the XL sheet"
#print (output.center(100, "^"))


#sheet1 = wb.get_sheet_by_name('Classifiers')
#sheet2 = wb.get_sheet_by_name('To be classified')

#for i in (range (16)):
#    sheet1.cell(row=(5+i), column=1).value = W1[i]

#for i in (range (16)):
#    for j in (range (6)):
#        sheet1.cell(row=(5+i), column=(5+j)).value = W2[i][j]

#for i in (range (50)):
#    sheet2.cell(row=(5+i), column=16).value = t1[i]
#    sheet2.cell(row=(5+i), column=17).value = t2[i]

#wb.save('Assignment_4_Data_and_Template.xlsx')
